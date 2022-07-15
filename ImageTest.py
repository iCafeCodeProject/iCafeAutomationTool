#!/user/bin/python3
# -*- coding -*-
# @Time    : 2022/5/5
# @Author  :
# @Site    : 
# @File    : ImageTest.py
# @Software: PyCharm
# @Description:


# 导入所需Package
import time
import cv2
import os
from PIL import ImageGrab
from PIL import Image
import numpy as np
import matplotlib.pyplot as plt
import openpyxl

wb = openpyxl.Workbook()
ws = wb.create_sheet("原神运行数据")
ws.cell(row = 1, column =1).value = "Image"
ws.cell(row = 1, column =2).value = "ImageVar"




# 方法一：截取图片并判断
start = time.time_ns()

k = 2
g = 2
t = 1000
thresholdmin = 100
thresholdmax = 3000

for x in range(t):
    image = ImageGrab.grab()
    imagenum = np.asarray(image)
    # plt.imshow(image)
    # plt.show()

    imsize = imagenum.shape
    imagenum = cv2.resize(imagenum, (int(imsize[1] / 2), int(imsize[0] / 2)), interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(imagenum, cv2.COLOR_BGR2GRAY)
    imageVar = cv2.Laplacian(gray, cv2.CV_64F).var()
    imageVar = int(imageVar)
    time.sleep(1)
    if imageVar < thresholdmin or imageVar > thresholdmax:
        # cv2.imshow('', image)
        # cv2.waitKey(0)
        # print(str(x) + '  imageVar: %d' % (imageVar))

        ws.cell(row=k, column=1).value = str(x)
        ws.cell(row=g, column=2).value = str(imageVar)
        k = k + 1
        g = g + 1
        wb.save("运行数据.xlsx")
        image.save("./imagefail/" + str(x) +"_"+ str(imageVar)+ ".png")


end = time.time_ns()
print('running spend: ' + str((end - start) / 1000 / 1000 / 1000) + 's')


# # 方法二：保存图片并判断
#
# t = 10
# dirpath = './image0/'
# thresholdmin = 300
# thresholdmax = 2500
#
# for x in range(t):
#     image = ImageGrab.grab()
#     image.save(dirpath + str(x) + ".png")
#
# if os.path.isdir(dirpath) == False:
#     raise AssertionError
#
# files = os.listdir(dirpath)
#
# for filename in files:
#
#     file = os.path.join(dirpath, filename)
#
#     image = cv2.imread(file)
#
#     imsize = image.shape
#
#     image = cv2.resize(image, (int(imsize[1]/2), int(imsize[0]/2)), interpolation=cv2.INTER_CUBIC)
#
#     # cv2.imshow('', image)
#     # cv2.waitKey(0)
#
#     gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
#
#     # cv2.imshow('', gray)
#     # cv2.waitKey(0)
#
#     imageVar = cv2.Laplacian(gray, cv2.CV_64F).var()
#
#     if imageVar < thresholdmin or imageVar > thresholdmax:
#         # cv2.imshow('', image)
#         # cv2.waitKey(0)
#         print(filename + '  imageVar: %d' % (imageVar))
#         cv2.imwrite("./imagefail/" + str(x) + ".png",image)
