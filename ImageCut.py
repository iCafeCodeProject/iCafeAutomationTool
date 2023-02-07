from cv2 import resize, cvtColor, Laplacian, CV_64F, COLOR_BGR2GRAY, INTER_CUBIC
import numpy as np
from PIL import ImageGrab
import time
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

threshold_min = 0
threshold_max = 100000


def image_processing(threshold_min, threshold_max, image):
    imagenum = np.asarray(image)
    imsize = imagenum.shape
    imagenum = resize(imagenum, (int(imsize[1] / 2), int(imsize[0] / 2)), interpolation=INTER_CUBIC)
    gray = cvtColor(imagenum, COLOR_BGR2GRAY)
    imageVar = Laplacian(gray, CV_64F).var()
    # if imageVar < threshold_min or imageVar > threshold_max:
    if threshold_min < imageVar < threshold_max:
        result = 1
    else:
        result = 0
    return result, imageVar

im_num = 0
run_path = './'+ str('TestDemo' + '_' + datetime.datetime.now().strftime('%m-%d-%H-%M-%S') + '\\')
os.mkdir(run_path)

wb = Workbook()
ws = wb.active
ws.title = "运行数据"
ws.cell(row=1, column=1).value = "ImageName"
ws.cell(row=1, column=2).value = "ImageTime"
ws.cell(row=1, column=3).value = "ImageVar"
exl_row = 2

while True:
    image = ImageGrab.grab()
    result, imageVar = image_processing(threshold_min, threshold_max, image)

    ws.cell(row=exl_row, column=1).value = str(im_num)
    ws.cell(row=exl_row, column=2).value = str(datetime.datetime.now().strftime('%H:%M:%S'))
    ws.cell(row=exl_row, column=3).value = str(int(imageVar))

    # fille = PatternFill('solid', fgColor='FF00FF')

    if result == 1:
        image.save(run_path + str(im_num) + ".png")

        # ws.cell(row=exl_row, column=1).fill = fille
        # ws.cell(row=exl_row, column=2).fill = fille
        # ws.cell(row=exl_row, column=3).fill = fille

    im_num = im_num + 1
    exl_row = exl_row + 1
    wb.save(run_path + "运行数据.xlsx")

    time.sleep(1)



