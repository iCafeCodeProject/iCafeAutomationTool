import cv2
from cv2 import resize, cvtColor, Laplacian, CV_64F, COLOR_BGR2GRAY, INTER_CUBIC
import numpy as np
from PIL import Image
import time
import os
import datetime
import threading
from openpyxl import Workbook


im_num = 0
wb = Workbook()
ws = wb.active
ws.title = "Run_data"
ws.cell(row=1, column=1).value = "ImageName"
ws.cell(row=1, column=2).value = "ImageTime"
ws.cell(row=1, column=3).value = "ImageVar"
exl_row = 2
image = 0
run_path = '0'
threshold_min = 0
threshold_max = 0
iqa_stop_flag = 0


def start_iqa():
    global iqa_timer, iqa_stop_flag
    if iqa_stop_flag:
        iqa_stop_flag = 0
        return 0
    image_processing()
    iqa_timer = threading.Timer(1, start_iqa)
    iqa_timer.start()


def image_processing():
    global im_num, exl_row
    imagenum = np.asarray(image)
    imsize = imagenum.shape
    imagenum = resize(imagenum, (int(imsize[1] / 2), int(imsize[0] / 2)), interpolation=INTER_CUBIC)
    gray = cvtColor(imagenum, COLOR_BGR2GRAY)
    imageVar = Laplacian(gray, CV_64F).var()
    # if imageVar < threshold_min or imageVar > threshold_max:

    ws.cell(row=exl_row, column=1).value = str(im_num)
    ws.cell(row=exl_row, column=2).value = str(datetime.datetime.now().strftime('%H:%M:%S'))
    ws.cell(row=exl_row, column=3).value = str(int(imageVar))

    if threshold_min < imageVar < threshold_max:
        im_num = im_num + 1
        exl_row = exl_row + 1

    else:
        cv2.imwrite(str(run_path + str(im_num) + ".png"), image)
        im_num = im_num + 1
        exl_row = exl_row + 1


def iqa_test(ui):
    global image, im_num, exl_row, threshold_max, threshold_min, iqa_stop_flag, run_path, iqa_timer
    threshold_max = int(ui.lineEdit_Threshold_Max.text())
    threshold_min = int(ui.lineEdit_Threshold_Min.text())
    run_path = './' + str('IQA_results' + '_' + datetime.datetime.now().strftime('%m-%d-%H-%M-%S') + '\\')
    os.mkdir(run_path)
    msg = str("[IQA]: IQA module start!...")
    ui.msg_print(msg)

    video = cv2.VideoCapture("image_corruption.mp4")
    status, frame = video.read()
    if status == 1:
        image = frame
        cv2.waitKey(10)
    iqa_timer = threading.Timer(1, start_iqa)
    iqa_timer.start()
    while True:
        status, frame = video.read()
        if status == 1:
            cv2.imshow("video", frame)
            image = frame
            cv2.waitKey(10)
        else:
            video.release()
            cv2.destroyAllWindows()
            im_num = 0
            exl_row = 2
            iqa_stop_flag = 1
            wb.save(run_path + "Run_data.xlsx")
            iqa_timer.join()
            msg = str("[IQA]: Running complete!...")
            ui.msg_print(msg)
            break
    return 0


iqa_timer = threading.Timer(1, start_iqa)
